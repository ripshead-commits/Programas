from fastapi import FastAPI, APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import io
import logging
import uuid
from pathlib import Path
from datetime import datetime, timezone
from typing import List, Dict, Any, Optional

from pydantic import BaseModel, Field
from openpyxl import load_workbook

from field_schemas import MODULES, get_module, list_modules

ROOT_DIR = Path(__file__).parent
TEMPLATES_DIR = ROOT_DIR / "templates"
load_dotenv(ROOT_DIR / ".env")

mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ["DB_NAME"]]

app = FastAPI(title="Plantillas FUR / FRG (ADRES)")
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)


# --------------------------------------------------------------------------- #
# Models
# --------------------------------------------------------------------------- #
class SubmissionCreate(BaseModel):
    data: Dict[str, Any]
    label: Optional[str] = None


class Submission(BaseModel):
    id: str
    module_id: str
    data: Dict[str, Any]
    label: Optional[str] = None
    created_at: str
    updated_at: str


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def _validate_payload(module_id: str, data: Dict[str, Any]) -> None:
    module = get_module(module_id)
    if not module:
        raise HTTPException(status_code=404, detail=f"Módulo '{module_id}' no encontrado")

    errors: List[str] = []
    for field in module["fields"]:
        key = field["key"]
        value = data.get(key)
        is_empty = value is None or (isinstance(value, str) and value.strip() == "")
        if field.get("required") and is_empty:
            errors.append(f"El campo '{field['label']}' es obligatorio.")
            continue
        if is_empty:
            continue
        # length
        max_len = field.get("length")
        if max_len and isinstance(value, (str, int, float)):
            if len(str(value)) > max_len:
                errors.append(f"'{field['label']}' excede la longitud máxima de {max_len} caracteres.")
        # type
        ftype = field["type"]
        if ftype == "number":
            try:
                float(str(value).replace(",", "."))
            except (TypeError, ValueError):
                errors.append(f"'{field['label']}' debe ser numérico.")
        if ftype == "select" and field.get("options"):
            allowed = [o["value"] for o in field["options"]]
            if str(value) not in allowed:
                errors.append(f"'{field['label']}' no es un valor válido.")
        if ftype == "date" and isinstance(value, str):
            try:
                datetime.strptime(value, "%Y-%m-%d")
            except ValueError:
                errors.append(f"'{field['label']}' debe tener formato AAAA-MM-DD.")
        if ftype == "time" and isinstance(value, str):
            try:
                datetime.strptime(value, "%H:%M")
            except ValueError:
                errors.append(f"'{field['label']}' debe tener formato HH:MM.")

    if errors:
        raise HTTPException(status_code=422, detail={"errors": errors})


def _fill_template(module_id: str, data: Dict[str, Any]) -> io.BytesIO:
    module = get_module(module_id)
    template_path = TEMPLATES_DIR / module["template_filename"]
    wb = load_workbook(template_path)
    ws = wb.active

    # Mapear nombre de columna (fila 1) a su índice
    headers: Dict[str, int] = {}
    for cell in ws[1]:
        if cell.value is not None:
            headers[str(cell.value).strip()] = cell.column

    # Escribir en fila 2 los valores recibidos
    for field in module["fields"]:
        key = field["key"]
        if key not in headers:
            continue
        col = headers[key]
        value = data.get(key, "")
        if value is None:
            value = ""
        if field["type"] == "number" and value != "":
            try:
                num = float(str(value).replace(",", "."))
                value = int(num) if num.is_integer() else num
            except (TypeError, ValueError):
                pass
        ws.cell(row=2, column=col, value=value)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# --------------------------------------------------------------------------- #
# Endpoints
# --------------------------------------------------------------------------- #
@api_router.get("/")
async def root():
    return {"message": "API Plantillas FUR / FRG (ADRES) operativa", "modules": list_modules()}


@api_router.get("/modules")
async def get_modules():
    return list_modules()


@api_router.get("/modules/{module_id}/schema")
async def get_schema(module_id: str):
    module = get_module(module_id)
    if not module:
        raise HTTPException(status_code=404, detail="Módulo no encontrado")

    # Agrupar campos por sección preservando orden
    sections: List[Dict[str, Any]] = []
    seen = {}
    for field in module["fields"]:
        sec_name = field.get("section") or "General"
        if sec_name not in seen:
            seen[sec_name] = len(sections)
            sections.append({"name": sec_name, "fields": []})
        sections[seen[sec_name]]["fields"].append(field)

    return {
        "id": module["id"],
        "name": module["name"],
        "description": module["description"],
        "sections": sections,
        "field_count": len(module["fields"]),
    }


@api_router.post("/modules/{module_id}/submissions")
async def create_submission(module_id: str, payload: SubmissionCreate):
    _validate_payload(module_id, payload.data)
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "id": str(uuid.uuid4()),
        "module_id": module_id.lower(),
        "data": payload.data,
        "label": payload.label,
        "created_at": now,
        "updated_at": now,
    }
    await db.submissions.insert_one(dict(doc))
    return doc


@api_router.get("/modules/{module_id}/submissions")
async def list_submissions(module_id: str):
    if not get_module(module_id):
        raise HTTPException(status_code=404, detail="Módulo no encontrado")
    cursor = db.submissions.find({"module_id": module_id.lower()}, {"_id": 0}).sort("created_at", -1)
    items = await cursor.to_list(500)
    return items


@api_router.get("/modules/{module_id}/submissions/{submission_id}")
async def get_submission(module_id: str, submission_id: str):
    doc = await db.submissions.find_one(
        {"id": submission_id, "module_id": module_id.lower()}, {"_id": 0}
    )
    if not doc:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    return doc


@api_router.delete("/modules/{module_id}/submissions/{submission_id}")
async def delete_submission(module_id: str, submission_id: str):
    res = await db.submissions.delete_one(
        {"id": submission_id, "module_id": module_id.lower()}
    )
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    return {"deleted": True, "id": submission_id}


@api_router.post("/modules/{module_id}/download")
async def download_template(module_id: str, payload: SubmissionCreate):
    module = get_module(module_id)
    if not module:
        raise HTTPException(status_code=404, detail="Módulo no encontrado")
    _validate_payload(module_id, payload.data)

    # Persistir como histórico
    now = datetime.now(timezone.utc).isoformat()
    await db.submissions.insert_one(
        {
            "id": str(uuid.uuid4()),
            "module_id": module_id.lower(),
            "data": payload.data,
            "label": payload.label,
            "created_at": now,
            "updated_at": now,
            "downloaded": True,
        }
    )

    buffer = _fill_template(module_id, payload.data)
    filename = module["download_filename"]
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
