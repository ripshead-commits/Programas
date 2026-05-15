"""Backend tests for Plantillas FUR/SER/FRG (ADRES)."""
import io
import os
import pytest
import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv("/app/frontend/.env")
BASE_URL = os.environ["REACT_APP_BACKEND_URL"].rstrip("/")
API = f"{BASE_URL}/api"

SER_VALID = {
    "NUM_FACTURA": "FE-TEST-001",
    "NIT_PRESTADOR": "900123456",
    "Tipo_de_servicio": "1",
    "Codigo_del_servicio": "S001",
    "Codificacion_CUPS": "990101",
    "Descripcion_del_servicio_o_elemento_reclamado": "Acetaminofen 500mg tableta",
    "Cantidad_de_servicios": "10",
    "Valor_unitario_facturado": "1000",
    "Valor_unitario_reclamado": "1000",
    "Valor_total_facturado": "10000",
    "Valor_total_reclamado": "10000",
}


@pytest.fixture(scope="module")
def session():
    s = requests.Session()
    s.headers.update({"Content-Type": "application/json"})
    return s


# --- Modules listing ---
class TestModules:
    def test_list_modules(self, session):
        r = session.get(f"{API}/modules")
        assert r.status_code == 200
        data = r.json()
        ids = {m["id"]: m["field_count"] for m in data}
        assert ids.get("fur") == 62
        assert ids.get("ser") == 13
        assert ids.get("frg") == 63

    @pytest.mark.parametrize("mid,count", [("fur", 62), ("ser", 13), ("frg", 63)])
    def test_schema(self, session, mid, count):
        r = session.get(f"{API}/modules/{mid}/schema")
        assert r.status_code == 200
        d = r.json()
        assert d["id"] == mid
        assert d["field_count"] == count
        assert isinstance(d["sections"], list) and len(d["sections"]) > 0
        # ensure fields grouped
        total = sum(len(s["fields"]) for s in d["sections"])
        assert total == count

    def test_schema_404(self, session):
        r = session.get(f"{API}/modules/xxx/schema")
        assert r.status_code == 404


# --- Validation ---
class TestValidation:
    def test_partial_data_422(self, session):
        r = session.post(f"{API}/modules/ser/submissions", json={"data": {"NUM_FACTURA": "X"}})
        assert r.status_code == 422
        det = r.json()["detail"]
        assert "errors" in det
        assert any("obligatorio" in e.lower() for e in det["errors"])

    def test_tipo_servicio_not_numeric_rejected(self, session):
        d = dict(SER_VALID)
        d["Tipo_de_servicio"] = "abc"
        r = session.post(f"{API}/modules/ser/submissions", json={"data": d})
        assert r.status_code == 422
        errs = r.json()["detail"]["errors"]
        assert any("válido" in e.lower() or "numérico" in e.lower() for e in errs)

    def test_tipo_servicio_invalid_option_rejected(self, session):
        d = dict(SER_VALID)
        d["Tipo_de_servicio"] = "99"
        r = session.post(f"{API}/modules/ser/submissions", json={"data": d})
        assert r.status_code == 422
        errs = r.json()["detail"]["errors"]
        assert any("válido" in e.lower() for e in errs)

    def test_date_format_validation(self, session):
        # FUR requires Fecha_de_ocurrencia_evento (date). Send bad date alongside other partial data
        bad = {
            "NIT_PRESTADOR": "900123456",
            "NUM_FACTURA": "F1",
            "Fecha_de_ocurrencia_evento": "31/12/2024",
        }
        r = session.post(f"{API}/modules/fur/submissions", json={"data": bad})
        assert r.status_code == 422
        errs = r.json()["detail"]["errors"]
        assert any("AAAA-MM-DD" in e for e in errs)


# --- CRUD SER ---
class TestSerCRUD:
    created_id = None

    def test_create_ser(self, session):
        r = session.post(f"{API}/modules/ser/submissions", json={"data": SER_VALID, "label": "TEST_ser_one"})
        assert r.status_code == 200, r.text
        body = r.json()
        assert "id" in body
        assert body["module_id"] == "ser"
        assert body["data"]["NUM_FACTURA"] == "FE-TEST-001"
        TestSerCRUD.created_id = body["id"]

    def test_list_ser(self, session):
        r = session.get(f"{API}/modules/ser/submissions")
        assert r.status_code == 200
        items = r.json()
        assert isinstance(items, list)
        assert any(i["id"] == TestSerCRUD.created_id for i in items)
        # sorted desc by created_at
        ts = [i["created_at"] for i in items]
        assert ts == sorted(ts, reverse=True)

    def test_get_ser(self, session):
        assert TestSerCRUD.created_id
        r = session.get(f"{API}/modules/ser/submissions/{TestSerCRUD.created_id}")
        assert r.status_code == 200
        assert r.json()["id"] == TestSerCRUD.created_id

    def test_delete_ser(self, session):
        assert TestSerCRUD.created_id
        r = session.delete(f"{API}/modules/ser/submissions/{TestSerCRUD.created_id}")
        assert r.status_code == 200
        assert r.json()["deleted"] is True
        # 404 after delete
        r2 = session.get(f"{API}/modules/ser/submissions/{TestSerCRUD.created_id}")
        assert r2.status_code == 404


# --- Download ---
class TestDownload:
    def test_download_ser_xlsx(self, session):
        r = session.post(f"{API}/modules/ser/download", json={"data": SER_VALID, "label": "TEST_dl"})
        assert r.status_code == 200
        ct = r.headers.get("content-type", "")
        assert "spreadsheetml" in ct
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        headers = {str(c.value).strip(): c.column for c in ws[1] if c.value}
        # row 2 should contain submitted values
        for key in ["NUM_FACTURA", "NIT_PRESTADOR", "Tipo_de_servicio", "Codificacion_CUPS"]:
            assert key in headers, f"Column {key} not in template"
            cell = ws.cell(row=2, column=headers[key]).value
            assert str(cell) == str(SER_VALID[key]) or (
                isinstance(cell, (int, float)) and str(int(cell)) == str(SER_VALID[key])
            ), f"{key} mismatch: got {cell!r}"

    def test_download_partial_fails(self, session):
        r = session.post(f"{API}/modules/ser/download", json={"data": {"NUM_FACTURA": "X"}})
        assert r.status_code == 422
