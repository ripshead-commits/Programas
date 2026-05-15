"""
Generación del XLSX usando la plantilla original (preserva formato).
"""
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from field_schemas import get_module


def _templates_dir() -> Path:
    """Carpeta `templates/` ubicada junto al .exe o junto al script."""
    if getattr(sys, "frozen", False):
        # PyInstaller --add-data coloca templates dentro de _MEIPASS
        base = Path(getattr(sys, "_MEIPASS", Path(sys.executable).parent))
    else:
        base = Path(__file__).parent
    return base / "templates"


def validate_data(module_id: str, data: dict) -> list[str]:
    """Valida `data` contra el esquema del módulo. Devuelve lista de errores."""
    module = get_module(module_id)
    if not module:
        return [f"Módulo '{module_id}' no encontrado."]

    errors: list[str] = []
    for field in module["fields"]:
        key = field["key"]
        value = data.get(key)
        empty = value is None or (isinstance(value, str) and value.strip() == "")

        if field.get("required") and empty:
            errors.append(f"El campo '{field['label']}' es obligatorio.")
            continue
        if empty:
            continue

        max_len = field.get("length")
        if max_len and len(str(value)) > max_len:
            errors.append(
                f"'{field['label']}' excede la longitud máxima de {max_len} caracteres."
            )

        ftype = field["type"]
        if ftype == "number":
            try:
                float(str(value).replace(",", "."))
            except (TypeError, ValueError):
                errors.append(f"'{field['label']}' debe ser numérico.")
        if ftype == "select" and field.get("options"):
            allowed = [v for v, _ in field["options"]]
            if str(value) not in allowed:
                errors.append(f"'{field['label']}' no es un valor válido.")
        if ftype == "date" and isinstance(value, str) and value.strip():
            try:
                datetime.strptime(value, "%Y-%m-%d")
            except ValueError:
                errors.append(f"'{field['label']}' debe tener formato AAAA-MM-DD.")
        if ftype == "time" and isinstance(value, str) and value.strip():
            try:
                datetime.strptime(value, "%H:%M")
            except ValueError:
                errors.append(f"'{field['label']}' debe tener formato HH:MM.")
    return errors


def fill_template(module_id: str, data: dict, output_path: Path) -> Path:
    """Carga la plantilla, escribe los datos en fila 2 y la guarda en `output_path`."""
    module = get_module(module_id)
    if not module:
        raise ValueError(f"Módulo '{module_id}' no encontrado")

    template_path = _templates_dir() / module["template_filename"]
    if not template_path.exists():
        raise FileNotFoundError(
            f"No se encontró la plantilla: {template_path}.\n"
            f"Asegúrate de que la carpeta 'templates/' acompañe al ejecutable."
        )

    wb = load_workbook(template_path)
    ws = wb.active

    # Mapear nombre de columna (fila 1) a su índice
    headers: dict[str, int] = {}
    for cell in ws[1]:
        if cell.value is not None:
            headers[str(cell.value).strip()] = cell.column

    for field in module["fields"]:
        key = field["key"]
        if key not in headers:
            continue
        col = headers[key]
        value = data.get(key, "")
        if value is None:
            value = ""
        if field["type"] == "number" and value != "":
            try:
                num = float(str(value).replace(",", "."))
                value = int(num) if num.is_integer() else num
            except (TypeError, ValueError):
                pass
        ws.cell(row=2, column=col, value=value)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
